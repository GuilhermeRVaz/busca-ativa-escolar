import logging
import re
import time
import unicodedata
from pathlib import Path
from typing import Optional

import gspread
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment

from config import Settings, get_settings
from whatsapp_bot import WhatsAppLinkBuilder


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
logger = logging.getLogger(__name__)


class ActiveSchoolSearchProcessor:
    def __init__(self, settings: Optional[Settings] = None) -> None:
        self.settings = settings or get_settings()
        self.link_builder = WhatsAppLinkBuilder(
            self.settings.whatsapp_message_template,
        )

    def load_absence_report(self, report_path: Optional[Path] = None) -> pd.DataFrame:
        path = Path(report_path or self.settings.consolidated_report_path)
        logger.info("Lendo relatorio consolidado: %s", path)

        raw_df = pd.read_excel(path, header=None)
        header_row = self._find_absence_header_row(raw_df)
        df = pd.read_excel(path, header=header_row)
        df = df.rename(columns={"Nome": "student_name", "RA": "ra_raw"})
        df = df.dropna(subset=["student_name", "ra_raw"], how="any")

        day_columns = [
            column
            for column in df.columns
            if isinstance(column, (int, float))
            or (isinstance(column, str) and str(column).strip().isdigit())
        ]

        prepared = pd.DataFrame()
        first_column = df.columns[0] if len(df.columns) > 0 else None
        if "Turma" in df.columns:
            class_series = df["Turma"]
        elif first_column not in {"N°", "Nome", "RA"}:
            class_series = df[first_column]
        else:
            class_series = pd.Series([""] * len(df), index=df.index)
        prepared["class_name"] = class_series.fillna("").astype(str).str.strip()
        prepared["student_name"] = df["student_name"].astype(str).str.strip()
        prepared["ra_raw"] = df["ra_raw"].astype(str).str.strip()
        prepared["ra_base"] = prepared["ra_raw"].apply(self.extract_ra_base)
        prepared["ra_digit"] = prepared["ra_raw"].apply(self.extract_ra_digit)
        prepared["ra_key"] = prepared.apply(
            lambda row: self.build_ra_key(row["ra_base"], row["ra_digit"]),
            axis=1,
        )
        normalized_days = df[day_columns].apply(
            lambda column: column.map(self._absence_cell_to_int),
        )
        prepared["total_absences"] = normalized_days.sum(axis=1).astype(int)
        prepared["absence_days_with_records"] = normalized_days.gt(0).sum(axis=1).astype(int)

        def _extract_absence_days(row: pd.Series) -> str:
            return ", ".join(
                str(day)
                for day in day_columns
                if row.get(day, 0) > 0
            )

        prepared["absence_days"] = normalized_days.apply(_extract_absence_days, axis=1)

        prepared = prepared[prepared["ra_key"].notna()].copy()
        prepared = prepared[prepared["absence_days_with_records"] >= 2].copy()

        logger.info("Relatorio processado com %s aluno(s) com >=2 faltas.", len(prepared))
        return prepared

    def load_contacts_from_google_sheet(self) -> pd.DataFrame:
        if not self.settings.google_sheet_url:
            raise ValueError("Defina GOOGLE_SHEET_URL no arquivo .env.")

        credentials_file = self.settings.google_service_account_file
        if not credentials_file.exists():
            raise FileNotFoundError(
                f"Arquivo de credenciais nao encontrado: {credentials_file}",
            )

        client = self._connect_gspread_with_retry(credentials_file)
        workbook = client.open_by_url(self.settings.google_sheet_url)

        worksheet_setting = self.settings.google_sheet_worksheet.strip()
        if worksheet_setting and worksheet_setting != "*":
            tab_names = [t.strip() for t in worksheet_setting.split(",")]
        else:
            tab_names = [ws.title for ws in workbook.worksheets()]

        logger.info("Lendo Google Sheets: abas %s", tab_names)

        all_records: list[pd.DataFrame] = []
        for tab in tab_names:
            try:
                worksheet = workbook.worksheet(tab)
                records = worksheet.get_all_records()
                if records:
                    df_tab = pd.DataFrame(records)
                    df_tab["_tab"] = tab
                    all_records.append(df_tab)
                    logger.info("Aba '%s': %s registro(s).", tab, len(records))
                else:
                    logger.warning("Aba '%s' esta vazia - ignorada.", tab)
            except Exception as exc:
                logger.warning("Erro ao ler aba '%s': %s - ignorada.", tab, exc)

        if not all_records:
            raise ValueError("Nenhuma aba do Google Sheets continha dados validos.")

        contacts_df = pd.concat(all_records, ignore_index=True)
        logger.info("Total de contatos carregados: %s", len(contacts_df))
        return self.prepare_contacts_dataframe(contacts_df)

    def _connect_gspread_with_retry(
        self,
        credentials_file: Path,
        max_attempts: int = 3,
    ) -> gspread.Client:
        for attempt in range(1, max_attempts + 1):
            try:
                logger.info(
                    "Conectando ao Google Sheets (tentativa %s/%s)...",
                    attempt,
                    max_attempts,
                )
                return gspread.service_account(filename=str(credentials_file))
            except Exception as exc:
                logger.warning("Falha na conexao com Google Sheets: %s", exc)
                if attempt == max_attempts:
                    raise

                wait_time = attempt * 3
                logger.info("Tentando novamente em %s segundos...", wait_time)
                time.sleep(wait_time)

        raise RuntimeError("Nao foi possivel conectar ao Google Sheets.")

    def prepare_contacts_dataframe(self, contacts_df: pd.DataFrame) -> pd.DataFrame:
        df = contacts_df.copy()
        renamed_columns = {column: self._normalize_column_name(column) for column in df.columns}
        df = df.rename(columns=renamed_columns)

        situacao_column = self._pick_column(df, ["situacao"])
        if situacao_column:
            situacoes_excluidas = {"TRAN", "BXTR"}
            mask_excluidos = df[situacao_column].astype(str).str.strip().str.upper().isin(situacoes_excluidas)
            n_excluidos = mask_excluidos.sum()
            if n_excluidos:
                logger.info("Excluindo %s aluno(s) com situacao TRAN/BXTR.", n_excluidos)
            df = df[~mask_excluidos].copy()

        ra_column = self._pick_column(df, ["ra"])
        ra_digit_column = self._pick_column(df, ["dig_ra", "digito_ra"])
        student_name_column = self._pick_column(
            df,
            ["nome_do_aluno", "nome_aluno", "aluno", "nome"],
        )

        if not ra_column:
            raise KeyError(
                "A planilha precisa ter, no minimo, a coluna RA.",
            )

        contact_slots = self._extract_contact_slots(df)
        if not contact_slots:
            raise KeyError(
                "A planilha precisa ter pelo menos uma combinacao de nome do responsavel e telefone.",
            )

        prepared = pd.DataFrame(index=df.index)
        prepared["ra_base"] = df[ra_column].apply(self.extract_ra_base)
        prepared["ra_digit"] = (
            df[ra_digit_column].astype(str).str.strip().str.upper()
            if ra_digit_column
            else ""
        )
        prepared["ra_key"] = prepared.apply(
            lambda row: self.build_ra_key(row["ra_base"], row["ra_digit"]),
            axis=1,
        )
        prepared["contact_student_name"] = (
            df[student_name_column].astype(str).str.strip()
            if student_name_column
            else ""
        )

        for index in range(1, 4):
            prepared[f"parent_name_{index}"] = ""
            prepared[f"phone_raw_{index}"] = ""
            prepared[f"phone_sanitized_{index}"] = ""

        for index, slot in enumerate(contact_slots, start=1):
            if index > 3:
                break
            prepared[f"parent_name_{index}"] = df[slot["name"]].fillna("").astype(str).str.strip()
            prepared[f"phone_raw_{index}"] = df[slot["phone"]].fillna("").astype(str).str.strip()
            prepared[f"phone_sanitized_{index}"] = prepared[f"phone_raw_{index}"].apply(
                self.sanitize_phone_number,
            )

        def _first_non_empty(series: pd.Series) -> str:
            for value in series:
                if pd.notna(value):
                    text = str(value).strip()
                    if text:
                        return text
            return ""

        aggregate_map: dict[str, object] = {
            "ra_base": "first",
            "ra_digit": "first",
            "contact_student_name": _first_non_empty,
        }
        for index in range(1, 4):
            aggregate_map[f"parent_name_{index}"] = _first_non_empty
            aggregate_map[f"phone_raw_{index}"] = _first_non_empty
            aggregate_map[f"phone_sanitized_{index}"] = _first_non_empty

        prepared = prepared.dropna(subset=["ra_key"]).groupby("ra_key", as_index=False).agg(aggregate_map)
        prepared["parent_name"] = prepared["parent_name_1"].replace("", "Responsavel")
        prepared["phone_raw"] = prepared["phone_raw_1"]
        prepared["phone_sanitized"] = prepared["phone_sanitized_1"]
        prepared["contact_slot"] = "responsavel_1"
        prepared["contact_found"] = True

        logger.info("Contatos validos apos limpeza: %s", len(prepared))
        return prepared

    def merge_absences_with_contacts(
        self,
        absence_df: pd.DataFrame,
        contacts_df: pd.DataFrame,
    ) -> pd.DataFrame:
        logger.info("Executando merge por RA normalizado.")
        merged = absence_df.merge(
            contacts_df,
            how="left",
            on="ra_key",
            suffixes=("", "_contact"),
        )
        merged["parent_name"] = merged["parent_name"].fillna("Responsavel")
        merged["phone_sanitized"] = merged["phone_sanitized"].fillna("")
        if "absence_days" not in merged.columns:
            merged["absence_days"] = ""
        merged["absence_days"] = merged["absence_days"].fillna("").astype(str)
        merged["contact_status"] = merged.apply(self._build_contact_status, axis=1)
        merged["whatsapp_message"] = merged.apply(
            lambda row: self.link_builder.build_message(
                row["parent_name"],
                row["student_name"],
                row["absence_days"],
            ),
            axis=1,
        )
        merged["whatsapp_link"] = merged.apply(
            lambda row: self.link_builder.build_chat_link(row["phone_sanitized"])
            if row["phone_sanitized"]
            else "",
            axis=1,
        )
        return merged

    def export_ready_to_send(
        self,
        merged_df: pd.DataFrame,
        output_path: Optional[Path] = None,
    ) -> Path:
        path = Path(output_path or self.settings.ready_to_send_output_path)
        logger.info("Salvando planilha final com multiplas abas: %s", path)

        if "contact_slot" not in merged_df.columns:
            logger.warning("Coluna 'contact_slot' nao encontrada. Salvando apenas aba unica.")
            merged_df.to_excel(path, index=False)
            return path

        hidden_columns = {
            "parent_name_2",
            "phone_raw_2",
            "phone_sanitized_2",
            "parent_name_3",
            "phone_raw_3",
            "phone_sanitized_3",
        }
        output_columns = [column for column in merged_df.columns if column not in hidden_columns]
        todos_df = merged_df[output_columns].copy()

        def _build_contact_sheet(slot_index: int) -> pd.DataFrame:
            parent_column = f"parent_name_{slot_index}"
            phone_raw_column = f"phone_raw_{slot_index}"
            phone_column = f"phone_sanitized_{slot_index}"
            required_columns = {parent_column, phone_raw_column, phone_column}

            if not required_columns.issubset(set(merged_df.columns)):
                return pd.DataFrame(columns=output_columns)

            slot_df = merged_df.copy()
            slot_df["parent_name"] = slot_df[parent_column].fillna("").replace("", "Responsavel")
            slot_df["phone_raw"] = slot_df[phone_raw_column].fillna("")
            slot_df["phone_sanitized"] = slot_df[phone_column].fillna("")
            slot_df["contact_slot"] = f"responsavel_{slot_index}"
            slot_df["contact_status"] = slot_df.apply(self._build_contact_status, axis=1)
            slot_df["whatsapp_message"] = slot_df.apply(
                lambda row: self.link_builder.build_message(
                    row["parent_name"],
                    row["student_name"],
                    row["absence_days"],
                ),
                axis=1,
            )
            slot_df["whatsapp_link"] = slot_df.apply(
                lambda row: self.link_builder.build_chat_link(row["phone_sanitized"])
                if row["phone_sanitized"]
                else "",
                axis=1,
            )
            return slot_df.loc[slot_df["phone_sanitized"].ne(""), output_columns].copy()

        contato_2 = _build_contact_sheet(2)
        contato_3 = _build_contact_sheet(3)

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            todos_df.to_excel(writer, sheet_name="Todos", index=False)

            if not contato_2.empty:
                contato_2.to_excel(writer, sheet_name="Contato_2", index=False)

            if not contato_3.empty:
                contato_3.to_excel(writer, sheet_name="Contato_3", index=False)

        workbook = openpyxl.load_workbook(path)
        for sheet_name in workbook.sheetnames:
            self._format_whatsapp_columns(workbook[sheet_name])
        workbook.save(path)

        return path

    def run(self) -> pd.DataFrame:
        absence_df = self.load_absence_report()
        contacts_df = self.load_contacts_from_google_sheet()
        merged_df = self.merge_absences_with_contacts(absence_df, contacts_df)
        self.export_ready_to_send(merged_df)
        logger.info("Pipeline da Fase 2 concluido.")
        return merged_df

    @staticmethod
    def _find_absence_header_row(df: pd.DataFrame) -> int:
        for index, row in df.iterrows():
            values = [str(value).strip().upper() for value in row.tolist() if pd.notna(value)]
            if {"N°", "NOME", "RA"}.issubset(set(values)):
                return index
        raise ValueError("Nao foi possivel localizar a linha de cabecalho do relatorio.")

    @staticmethod
    def _absence_cell_to_int(value: object) -> int:
        digits = re.sub(r"\D", "", str(value or ""))
        return int(digits) if digits else 0

    @staticmethod
    def extract_ra_base(value: object) -> Optional[str]:
        text = str(value or "").upper()
        match = re.search(r"(\d+)\s*-\s*([\dX])", text)
        if match:
            return match.group(1).lstrip("0") or "0"

        digits = re.sub(r"\D", "", text)
        if not digits:
            return None
        return digits.lstrip("0") or "0"

    @staticmethod
    def extract_ra_digit(value: object) -> str:
        text = str(value or "").upper()
        match = re.search(r"(\d+)\s*-\s*([\dX])", text)
        if match:
            return match.group(2)
        return ""

    @staticmethod
    def build_ra_key(ra_base: Optional[str], ra_digit: object) -> Optional[str]:
        if not ra_base:
            return None
        digit = str(ra_digit or "").strip().upper()
        return f"{ra_base}-{digit}" if digit else ra_base

    def sanitize_phone_number(self, value: object) -> str:
        digits = re.sub(r"\D", "", str(value or ""))
        if not digits:
            return ""

        country_code = self.settings.default_country_code
        default_ddd = self.settings.default_ddd

        if digits.startswith(country_code) and len(digits) in {12, 13}:
            return digits
        if len(digits) in {10, 11}:
            return f"{country_code}{digits}"
        if len(digits) in {8, 9}:
            return f"{country_code}{default_ddd}{digits}"
        if digits.startswith("0") and len(digits[1:]) in {10, 11}:
            return f"{country_code}{digits[1:]}"
        return ""

    @staticmethod
    def _normalize_column_name(value: object) -> str:
        text = str(value or "").strip().lower()
        text = "".join(
            char
            for char in unicodedata.normalize("NFKD", text)
            if not unicodedata.combining(char)
        )
        text = re.sub(r"[^a-z0-9]+", "_", text)
        return text.strip("_")

    @staticmethod
    def _pick_column(df: pd.DataFrame, candidates: list[str]) -> Optional[str]:
        for candidate in candidates:
            if candidate in df.columns:
                return candidate
        return None

    @staticmethod
    def _extract_contact_slots(df: pd.DataFrame) -> list[dict[str, str]]:
        slots: list[dict[str, str]] = []
        candidate_pairs = [
            ("responsavel_1", "telefone_1"),
            ("responsavel_2", "telefone_2"),
            ("responsavel_3", "telefone_3"),
            ("responsavel", "telefone"),
            ("nome_responsavel", "telefone_1"),
            ("nome_respons_vel", "telefone_1"),
            ("nome_responsavel", "telefone1"),
            ("nome_respons_vel", "telefone1"),
            ("nome_responsavel_2", "telefone_2"),
            ("nome_respons_vel_2", "telefone_2"),
            ("nome_responsavel_2", "telefone2"),
            ("nome_respons_vel_2", "telefone2"),
            ("nome_responsavel_3", "telefone_3"),
            ("nome_respons_vel_3", "telefone_3"),
            ("nome_responsavel_3", "telefone3"),
            ("nome_respons_vel_3", "telefone3"),
        ]

        for name_column, phone_column in candidate_pairs:
            if name_column in df.columns and phone_column in df.columns:
                slots.append({"name": name_column, "phone": phone_column})

        return slots

    @staticmethod
    def _build_contact_status(row: pd.Series) -> str:
        if pd.isna(row.get("contact_found")):
            return "RA nao encontrado na planilha de contatos"
        if not row.get("phone_sanitized"):
            return "Contato encontrado sem telefone valido"
        return "Pronto para envio"

    @staticmethod
    def _format_whatsapp_columns(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        header_map = {
            str(cell.value).strip(): index
            for index, cell in enumerate(worksheet[1], start=1)
            if cell.value
        }
        message_column = header_map.get("whatsapp_message")
        link_column = header_map.get("whatsapp_link")

        if message_column:
            worksheet.column_dimensions[
                openpyxl.utils.get_column_letter(message_column)
            ].width = 70
            for row in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row, column=message_column).alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

        if link_column:
            worksheet.column_dimensions[
                openpyxl.utils.get_column_letter(link_column)
            ].width = 18
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=link_column)
                url = str(cell.value or "").strip()
                if not url:
                    continue
                cell.value = "Abrir WhatsApp"
                cell.hyperlink = url
                cell.style = "Hyperlink"


if __name__ == "__main__":
    processor = ActiveSchoolSearchProcessor()
    processor.run()
