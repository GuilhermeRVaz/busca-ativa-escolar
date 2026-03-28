import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
GREEN_FILL = PatternFill("solid", fgColor="70AD47")
YELLOW_FILL = PatternFill("solid", fgColor="FFC000")
RED_FILL = PatternFill("solid", fgColor="C00000")
ORANGE_FILL = PatternFill("solid", fgColor="ED7D31")
BLUE_FILL = PatternFill("solid", fgColor="5B9BD5")
LIGHT_FILL = PatternFill("solid", fgColor="D9EAF7")


class DashboardBuilder:
    def __init__(self, reports_dir: Path) -> None:
        self.reports_dir = reports_dir

    def build(self, output_path: Path | None = None) -> Path:
        campaigns_df = self._load_campaigns()
        operational_df = self._load_operational()
        retornos_df, justificativas_df, revisao_df = self._load_retorno_bases()
        auditoria_df = self._load_auditoria()

        output = output_path or self.reports_dir / "Dashboard_Busca_Ativa.xlsx"
        workbook = Workbook()
        workbook.remove(workbook.active)

        painel_ws = workbook.create_sheet("Painel")
        operacional_ws = workbook.create_sheet("Operacional")
        retornos_ws = workbook.create_sheet("Retornos")
        justificativas_ws = workbook.create_sheet("Justificativas")
        revisao_ws = workbook.create_sheet("Revisao")
        base_modelo_ws = workbook.create_sheet("Base_Modelo")

        base_campaigns_ws = workbook.create_sheet("Base_Campanhas")
        base_operacional_ws = workbook.create_sheet("Base_Operacional")
        base_retorno_ws = workbook.create_sheet("Base_Retornos")
        base_justificativas_ws = workbook.create_sheet("Base_Justificativas")
        base_revisao_ws = workbook.create_sheet("Base_Revisao")
        base_auditoria_ws = workbook.create_sheet("Base_Auditoria")

        for sheet in [
            base_campaigns_ws,
            base_operacional_ws,
            base_retorno_ws,
            base_justificativas_ws,
            base_revisao_ws,
            base_auditoria_ws,
        ]:
            sheet.sheet_state = "hidden"

        self._build_painel(painel_ws, campaigns_df, retornos_df, justificativas_df, revisao_df)
        self._build_operacional(operacional_ws, campaigns_df, operational_df, retornos_df)
        self._build_retornos(retornos_ws, retornos_df)
        self._build_justificativas(justificativas_ws, justificativas_df)
        self._build_revisao(revisao_ws, revisao_df)
        self._build_base_modelo(base_modelo_ws)

        self._write_dataframe(base_campaigns_ws, campaigns_df)
        self._write_dataframe(base_operacional_ws, operational_df)
        self._write_dataframe(base_retorno_ws, retornos_df)
        self._write_dataframe(base_justificativas_ws, justificativas_df)
        self._write_dataframe(base_revisao_ws, revisao_df)
        self._write_dataframe(base_auditoria_ws, auditoria_df)

        workbook.save(output)
        return output

    def _load_campaigns(self) -> pd.DataFrame:
        records: list[pd.DataFrame] = []
        for path in sorted(self.reports_dir.glob("Campanha*.xlsx")):
            if self._should_ignore(path.name):
                continue
            df = pd.read_excel(path, sheet_name="Campanha")
            if df.empty:
                continue
            df = df.copy()
            df["source_workbook"] = path.name
            df["tipo_campanha"] = "diaria" if path.stem.lower().startswith("campanha_diaria_") else "mensal"
            df["data_campanha"] = df.apply(
                lambda row: self._extract_campaign_date(self._safe_text(row.get("campaign_id")), self._safe_text(row.get("data_criacao"))),
                axis=1,
            )
            df["dia_referente_falta"] = df["campaign_id"].apply(self._extract_absence_day)
            df["turma_base"] = df["class_name"].apply(self._safe_text)
            df["status_envio_norm"] = df["status_envio"].apply(self._normalize_status)
            df["status_resposta_norm"] = df["status_resposta"].apply(self._normalize_status)
            df["enviado_flag"] = df["status_envio_norm"].eq("enviado")
            df["falha_flag"] = df["status_envio_norm"].eq("falha")
            df["numero_invalido_flag"] = df["status_resposta_norm"].eq("numero_invalido")
            records.append(df)
        return self._concat_or_empty(records)

    def _load_operational(self) -> pd.DataFrame:
        records: list[pd.DataFrame] = []
        for path in sorted(self.reports_dir.glob("Relatorio_Operacional_*.xlsx")):
            if self._should_ignore(path.name):
                continue
            df = pd.read_excel(path, sheet_name="Detalhes")
            if df.empty:
                continue
            df = df.copy()
            df["source_workbook"] = path.name
            df["tipo_campanha"] = "diaria" if "Campanha_Diaria_" in path.stem else "mensal"
            df["data_campanha"] = df["campaign_id"].apply(self._extract_campaign_date)
            df["status_envio_norm"] = df["status_envio"].apply(self._normalize_status)
            df["status_resposta_norm"] = df["status_resposta"].apply(self._normalize_status)
            records.append(df)
        return self._concat_or_empty(records)

    def _load_retorno_bases(self) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        retorno_frames: list[pd.DataFrame] = []
        justificativas_frames: list[pd.DataFrame] = []
        revisao_frames: list[pd.DataFrame] = []

        for path in sorted(self.reports_dir.glob("Relatorio_de_Retornos_*.xlsx")):
            if self._should_ignore(path.name):
                continue
            retorno_frames.extend(self._load_retorno_file(path))
            justificativas_df = pd.read_excel(path, sheet_name="Justificativas")
            if not justificativas_df.empty:
                justificativas_df = justificativas_df.copy()
                justificativas_df["source_workbook"] = path.name
                justificativas_df["data_campanha"] = justificativas_df["campaign_id"].apply(self._extract_campaign_date)
                justificativas_frames.append(justificativas_df)
            revisao_df = pd.read_excel(path, sheet_name="Revisar")
            if not revisao_df.empty:
                revisao_df = revisao_df.copy()
                revisao_df["source_workbook"] = path.name
                revisao_df["data_campanha"] = revisao_df["campaign_id"].apply(self._extract_campaign_date)
                revisao_frames.append(revisao_df)

        return (
            self._concat_or_empty(retorno_frames),
            self._concat_or_empty(justificativas_frames),
            self._concat_or_empty(revisao_frames),
        )

    def _load_retorno_file(self, path: Path) -> list[pd.DataFrame]:
        frames: list[pd.DataFrame] = []
        mapping = {
            "Respondidos": "respondido",
            "Sem_Retorno": "sem_retorno",
            "Nao_Recontatar": "nao_recontatar",
        }
        for sheet_name, bucket in mapping.items():
            df = pd.read_excel(path, sheet_name=sheet_name)
            if df.empty:
                continue
            df = df.copy()
            df["retorno_bucket"] = bucket
            df["source_workbook"] = path.name
            df["data_campanha"] = df["campaign_id"].apply(self._extract_campaign_date)
            if "class_name" in df.columns:
                df["turma_base"] = df["class_name"].apply(self._safe_text)
            if "reason_category_suggested" in df.columns:
                df["reason_category_suggested"] = df["reason_category_suggested"].fillna("").astype(str)
            frames.append(df)
        return frames

    def _load_auditoria(self) -> pd.DataFrame:
        records: list[pd.DataFrame] = []
        ledger_map = {
            "Campaign_Ledger.xlsx": "mensal",
            "Daily_Campaign_Ledger.xlsx": "diaria",
        }
        for filename, ledger_type in ledger_map.items():
            path = self.reports_dir / filename
            if not path.exists():
                continue
            df = pd.read_excel(path, sheet_name="Historico")
            if df.empty:
                continue
            df = df.copy()
            df["ledger_tipo"] = ledger_type
            df["source_workbook"] = filename
            records.append(df)
        return self._concat_or_empty(records)

    def _build_painel(
        self,
        worksheet,
        campaigns_df: pd.DataFrame,
        retornos_df: pd.DataFrame,
        justificativas_df: pd.DataFrame,
        revisao_df: pd.DataFrame,
    ) -> None:
        self._set_title(worksheet, "Dashboard Busca Ativa")
        worksheet["A3"] = "Atualizacao"
        worksheet["B3"] = pd.Timestamp.now().strftime("%d/%m/%Y %H:%M")
        worksheet["D3"] = "Filtro manual"
        worksheet["E3"] = "Use o autofiltro nas abas Operacional, Retornos, Justificativas e Revisao."

        total_contatos = int(len(campaigns_df))
        total_enviados = int(campaigns_df.get("enviado_flag", pd.Series(dtype=bool)).sum()) if not campaigns_df.empty else 0
        total_falhas = int(campaigns_df.get("falha_flag", pd.Series(dtype=bool)).sum()) if not campaigns_df.empty else 0
        total_invalidos = int(campaigns_df.get("numero_invalido_flag", pd.Series(dtype=bool)).sum()) if not campaigns_df.empty else 0
        total_respondidos = int(retornos_df["retorno_bucket"].eq("respondido").sum()) if not retornos_df.empty else 0
        total_sem_retorno = int(retornos_df["retorno_bucket"].eq("sem_retorno").sum()) if not retornos_df.empty else 0
        total_revisao = int(len(revisao_df))
        taxa_resposta = round((total_respondidos / total_enviados) * 100, 2) if total_enviados else 0.0

        kpis = [
            ("Total Contatos", total_contatos, BLUE_FILL),
            ("Enviados", total_enviados, GREEN_FILL),
            ("Falhas", total_falhas, RED_FILL),
            ("Numeros Invalidos", total_invalidos, ORANGE_FILL),
            ("Respondidos", total_respondidos, GREEN_FILL),
            ("Sem Retorno", total_sem_retorno, YELLOW_FILL),
            ("Para Revisao", total_revisao, ORANGE_FILL),
            ("Taxa Resposta %", taxa_resposta, BLUE_FILL),
        ]
        self._write_kpi_row(worksheet, 5, kpis)

        envios_por_data = self._safe_group_count(campaigns_df, "data_campanha", "enviado_flag", truthy=True, label="total_enviados")
        resposta_por_turma = self._group_retornos_por_turma(retornos_df)
        motivos_df = self._safe_group_simple(justificativas_df, "reason_category_suggested", "quantidade")
        funil_df = pd.DataFrame(
            [
                {"etapa": "Contatos", "quantidade": total_contatos},
                {"etapa": "Enviados", "quantidade": total_enviados},
                {"etapa": "Respondidos", "quantidade": total_respondidos},
                {"etapa": "Justificados", "quantidade": int(len(justificativas_df))},
            ]
        )

        self._write_dataframe(worksheet, envios_por_data, start_row=10, start_col=1)
        self._write_dataframe(worksheet, resposta_por_turma, start_row=10, start_col=5)
        self._write_dataframe(worksheet, motivos_df, start_row=10, start_col=10)
        self._write_dataframe(worksheet, funil_df, start_row=10, start_col=14)

        self._add_line_chart(worksheet, "Envios por Data", 10, 1, 10 + len(envios_por_data), "P1")
        self._add_bar_chart(worksheet, "Resposta por Turma", 10, 5, 10 + len(resposta_por_turma), "P18")
        self._add_pie_chart(worksheet, "Justificativas", 10, 10, 10 + len(motivos_df), "Z1")
        self._add_bar_chart(worksheet, "Funil Busca Ativa", 10, 14, 10 + len(funil_df), "Z18")

        worksheet.freeze_panes = "A5"

    def _build_operacional(self, worksheet, campaigns_df: pd.DataFrame, operational_df: pd.DataFrame, retornos_df: pd.DataFrame) -> None:
        self._set_title(worksheet, "Operacional")
        operacional_resumo = self._build_operacional_summary(campaigns_df, retornos_df)
        self._write_dataframe(worksheet, operacional_resumo, start_row=3, start_col=1)
        self._write_dataframe(worksheet, operational_df, start_row=3, start_col=8, enable_filter=True)
        worksheet.freeze_panes = "A3"

    def _build_retornos(self, worksheet, retornos_df: pd.DataFrame) -> None:
        self._set_title(worksheet, "Retornos")
        resumo = self._safe_group_simple(retornos_df, "retorno_bucket", "quantidade")
        por_campanha = self._safe_group_count(retornos_df, "campaign_id", "retorno_bucket", truthy=None, label="total_registros")
        self._write_dataframe(worksheet, resumo, start_row=3, start_col=1)
        self._write_dataframe(worksheet, por_campanha, start_row=3, start_col=5)
        self._write_dataframe(worksheet, retornos_df, start_row=12, start_col=1, enable_filter=True)
        worksheet.freeze_panes = "A12"

    def _build_justificativas(self, worksheet, justificativas_df: pd.DataFrame) -> None:
        self._set_title(worksheet, "Justificativas")
        resumo = self._safe_group_simple(justificativas_df, "reason_category_suggested", "quantidade")
        por_turma = self._safe_group_simple(justificativas_df, "class_name", "quantidade")
        self._write_dataframe(worksheet, resumo, start_row=3, start_col=1)
        self._write_dataframe(worksheet, por_turma, start_row=3, start_col=5)
        self._write_dataframe(worksheet, justificativas_df, start_row=12, start_col=1, enable_filter=True)
        self._add_bar_chart(worksheet, "Justificativas por Categoria", 3, 1, 3 + len(resumo), "L3")
        worksheet.freeze_panes = "A12"

    def _build_revisao(self, worksheet, revisao_df: pd.DataFrame) -> None:
        self._set_title(worksheet, "Revisao")
        resumo = self._safe_group_simple(revisao_df, "review_reason", "quantidade")
        self._write_dataframe(worksheet, resumo, start_row=3, start_col=1)
        self._write_dataframe(worksheet, revisao_df, start_row=10, start_col=1, enable_filter=True)
        self._add_bar_chart(worksheet, "Pendencias de Revisao", 3, 1, 3 + len(resumo), "J3")
        worksheet.freeze_panes = "A10"

    def _build_base_modelo(self, worksheet) -> None:
        self._set_title(worksheet, "Base Modelo")
        instructions = [
            "1. Atualize os relatorios com o fluxo Python atual.",
            "2. Rode: python dashboard_builder.py",
            "3. Abra Dashboard_Busca_Ativa.xlsx.",
            "4. Use os filtros nas abas visiveis para navegar pelos dados.",
            "5. As abas Base_* ficam ocultas e servem como base tecnica do painel.",
        ]
        for index, line in enumerate(instructions, start=3):
            worksheet[f"A{index}"] = line
        worksheet["A10"] = "Fontes esperadas"
        worksheet["A11"] = "Campanha_*.xlsx / Campanha_Diaria_*.xlsx"
        worksheet["A12"] = "Relatorio_Operacional_*.xlsx"
        worksheet["A13"] = "Relatorio_de_Retornos_*.xlsx"
        worksheet["A14"] = "Campaign_Ledger.xlsx / Daily_Campaign_Ledger.xlsx"
        worksheet.column_dimensions["A"].width = 85

    def _build_operacional_summary(self, campaigns_df: pd.DataFrame, retornos_df: pd.DataFrame) -> pd.DataFrame:
        if campaigns_df.empty:
            return pd.DataFrame(columns=["class_name", "total_contatos", "enviados", "falhas", "invalidos", "respondidos", "sem_retorno"])

        base = campaigns_df.copy()
        summary = (
            base.groupby("class_name", dropna=False)
            .agg(
                total_contatos=("student_name", "count"),
                enviados=("enviado_flag", "sum"),
                falhas=("falha_flag", "sum"),
                invalidos=("numero_invalido_flag", "sum"),
            )
            .reset_index()
        )
        if retornos_df.empty:
            summary["respondidos"] = 0
            summary["sem_retorno"] = 0
            return summary

        retorno_summary = (
            retornos_df.groupby(["class_name", "retorno_bucket"], dropna=False)
            .size()
            .unstack(fill_value=0)
            .reset_index()
        )
        retorno_summary = retorno_summary.rename(columns={"respondido": "respondidos", "sem_retorno": "sem_retorno"})
        for column in ["respondidos", "sem_retorno"]:
            if column not in retorno_summary.columns:
                retorno_summary[column] = 0
        return summary.merge(retorno_summary[["class_name", "respondidos", "sem_retorno"]], on="class_name", how="left").fillna(0)

    def _write_kpi_row(self, worksheet, row: int, kpis: list[tuple[str, object, PatternFill]]) -> None:
        start_col = 1
        for label, value, fill in kpis:
            header_cell = worksheet.cell(row=row, column=start_col)
            value_cell = worksheet.cell(row=row + 1, column=start_col)
            header_cell.value = label
            value_cell.value = value
            header_cell.fill = fill
            value_cell.fill = fill
            header_cell.font = Font(color="FFFFFF", bold=True)
            value_cell.font = Font(color="FFFFFF", bold=True, size=14)
            header_cell.alignment = Alignment(horizontal="center")
            value_cell.alignment = Alignment(horizontal="center")
            worksheet.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=start_col + 1)
            worksheet.merge_cells(start_row=row + 1, start_column=start_col, end_row=row + 2, end_column=start_col + 1)
            worksheet.column_dimensions[get_column_letter(start_col)].width = 18
            worksheet.column_dimensions[get_column_letter(start_col + 1)].width = 4
            start_col += 3

    def _set_title(self, worksheet, title: str) -> None:
        worksheet["A1"] = title
        worksheet["A1"].fill = TITLE_FILL
        worksheet["A1"].font = Font(color="FFFFFF", bold=True, size=16)
        worksheet["A1"].alignment = Alignment(horizontal="center")
        worksheet.merge_cells("A1:J1")

    def _write_dataframe(
        self,
        worksheet,
        dataframe: pd.DataFrame,
        start_row: int = 1,
        start_col: int = 1,
        enable_filter: bool = False,
    ) -> tuple[int, int]:
        if dataframe is None or dataframe.empty:
            dataframe = pd.DataFrame({"sem_dados": []})

        columns = list(dataframe.columns)
        for col_index, column_name in enumerate(columns, start=start_col):
            cell = worksheet.cell(row=start_row, column=col_index)
            cell.value = column_name
            cell.fill = LIGHT_FILL
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row_offset, row in enumerate(dataframe.itertuples(index=False), start=1):
            for col_offset, value in enumerate(row, start=0):
                data_cell = worksheet.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)
                if row_offset % 2 == 0:
                    data_cell.fill = PatternFill("solid", fgColor="F7FBFF")

        end_row = start_row + len(dataframe)
        end_col = start_col + len(columns) - 1
        if enable_filter and len(columns) > 0:
            worksheet.auto_filter.ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"

        self._autofit(worksheet, start_col, end_col)
        return end_row, end_col

    def _autofit(self, worksheet, start_col: int, end_col: int) -> None:
        for col_index in range(start_col, end_col + 1):
            letter = get_column_letter(col_index)
            max_len = 0
            for cell in worksheet[letter]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
            worksheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 36)

    def _add_line_chart(self, worksheet, title: str, start_row: int, start_col: int, end_row: int, anchor: str) -> None:
        if end_row <= start_row + 1:
            return
        chart = LineChart()
        chart.title = title
        chart.height = 8
        chart.width = 12
        data = Reference(worksheet, min_col=start_col + 1, min_row=start_row, max_row=end_row)
        cats = Reference(worksheet, min_col=start_col, min_row=start_row + 1, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        worksheet.add_chart(chart, anchor)

    def _add_bar_chart(self, worksheet, title: str, start_row: int, start_col: int, end_row: int, anchor: str) -> None:
        if end_row <= start_row + 1:
            return
        chart = BarChart()
        chart.title = title
        chart.height = 8
        chart.width = 12
        data = Reference(worksheet, min_col=start_col + 1, min_row=start_row, max_row=end_row)
        cats = Reference(worksheet, min_col=start_col, min_row=start_row + 1, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        worksheet.add_chart(chart, anchor)

    def _add_pie_chart(self, worksheet, title: str, start_row: int, start_col: int, end_row: int, anchor: str) -> None:
        if end_row <= start_row + 1:
            return
        chart = PieChart()
        chart.title = title
        chart.height = 8
        chart.width = 10
        data = Reference(worksheet, min_col=start_col + 1, min_row=start_row, max_row=end_row)
        labels = Reference(worksheet, min_col=start_col, min_row=start_row + 1, max_row=end_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        worksheet.add_chart(chart, anchor)

    def _group_retornos_por_turma(self, retornos_df: pd.DataFrame) -> pd.DataFrame:
        if retornos_df.empty:
            return pd.DataFrame(columns=["class_name", "respondidos"])
        responded = retornos_df.loc[retornos_df["retorno_bucket"].eq("respondido")].copy()
        return self._safe_group_simple(responded, "class_name", "respondidos")

    def _safe_group_count(self, dataframe: pd.DataFrame, group_col: str, value_col: str, truthy: bool | None, label: str) -> pd.DataFrame:
        if dataframe.empty or group_col not in dataframe.columns:
            return pd.DataFrame(columns=[group_col, label])
        df = dataframe.copy()
        if truthy is None:
            grouped = df.groupby(group_col, dropna=False).size().reset_index(name=label)
        else:
            grouped = (
                df.groupby(group_col, dropna=False)[value_col]
                .sum()
                .reset_index(name=label)
            )
        grouped[group_col] = grouped[group_col].fillna("").astype(str)
        return grouped.sort_values(group_col)

    def _safe_group_simple(self, dataframe: pd.DataFrame, group_col: str, label: str) -> pd.DataFrame:
        if dataframe.empty or group_col not in dataframe.columns:
            return pd.DataFrame(columns=[group_col, label])
        grouped = dataframe.copy()
        grouped[group_col] = grouped[group_col].fillna("nao_informado").astype(str)
        return grouped.groupby(group_col, dropna=False).size().reset_index(name=label).sort_values(label, ascending=False)

    def _concat_or_empty(self, frames: list[pd.DataFrame]) -> pd.DataFrame:
        return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()

    def _should_ignore(self, filename: str) -> bool:
        lowered = filename.lower()
        return any(marker in lowered for marker in ["backup_", "runtime_autosave", "teste"])

    def _extract_campaign_date(self, campaign_id: str, fallback: str = "") -> str:
        match = re.search(r"(20\d{2})_(\d{2})_(\d{2})", self._safe_text(campaign_id))
        if match:
            return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
        timestamp = pd.to_datetime(fallback, errors="coerce")
        return timestamp.strftime("%Y-%m-%d") if pd.notna(timestamp) else ""

    def _extract_absence_day(self, campaign_id: str) -> str:
        match = re.search(r"_dia_(\d{1,2})", self._safe_text(campaign_id))
        return match.group(1) if match else ""

    @staticmethod
    def _normalize_status(value: object) -> str:
        return DashboardBuilder._safe_text(value).strip().lower()

    @staticmethod
    def _safe_text(value: object) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and pd.isna(value):
            return ""
        return str(value).strip()


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Gera um dashboard Excel da Busca Ativa usando os relatorios da pasta relatorios.")
    parser.add_argument(
        "--reports-dir",
        default="relatorios",
        help="Pasta com as campanhas e relatorios gerados.",
    )
    parser.add_argument(
        "--output",
        help="Caminho do workbook Excel de saida.",
    )
    return parser


def main() -> None:
    args = build_argument_parser().parse_args()
    builder = DashboardBuilder(reports_dir=Path(args.reports_dir))
    output = builder.build(output_path=Path(args.output) if args.output else None)
    print(output)


if __name__ == "__main__":
    main()
